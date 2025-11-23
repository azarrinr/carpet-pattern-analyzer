import streamlit as st
from PIL import Image
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from sklearn.cluster import KMeans
import anthropic
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Page config
st.set_page_config(
    page_title="Carpet Pattern Analyzer",
    page_icon="🎨",
    layout="wide"
)

def extract_grid_colors(image, grid_width, grid_depth):
    """
    Divide image into grid and extract average color from each cell
    """
    # Convert image to numpy array
    img_array = np.array(image.convert('RGB'))
    img_height, img_width = img_array.shape[:2]
    
    # Calculate cell dimensions
    cell_width = img_width / grid_width
    cell_height = img_height / grid_depth
    
    # Store colors for each cell
    grid_colors = []
    
    for row in range(grid_depth):
        row_colors = []
        for col in range(grid_width):
            # Calculate cell boundaries
            x_start = int(col * cell_width)
            x_end = int((col + 1) * cell_width)
            y_start = int(row * cell_height)
            y_end = int((row + 1) * cell_height)
            
            # Extract cell from image
            cell = img_array[y_start:y_end, x_start:x_end]
            
            # Calculate average color
            avg_color = cell.mean(axis=(0, 1))
            row_colors.append(avg_color / 255.0)  # Normalize to 0-1 range
        
        grid_colors.append(row_colors)
    
    return np.array(grid_colors)

def quantize_colors(grid_colors, n_colors):
    """
    Reduce the number of colors using K-means clustering
    """
    original_shape = grid_colors.shape
    
    # Reshape to list of colors
    colors_flat = grid_colors.reshape(-1, 3)
    
    # Apply K-means clustering
    kmeans = KMeans(n_clusters=n_colors, random_state=42, n_init=10)
    kmeans.fit(colors_flat)
    
    # Replace each color with its cluster center
    quantized_colors = kmeans.cluster_centers_[kmeans.labels_]
    
    # Reshape back to grid
    quantized_grid = quantized_colors.reshape(original_shape)
    
    # Return both the quantized grid and the unique palette
    unique_colors = kmeans.cluster_centers_
    
    return quantized_grid, unique_colors

def plot_color_grid(grid_colors):
    """
    Create a visual grid showing the colors
    """
    grid_depth, grid_width = grid_colors.shape[:2]
    
    # Create figure
    fig, ax = plt.subplots(figsize=(12, 12 * grid_depth / grid_width))
    ax.set_xlim(0, grid_width)
    ax.set_ylim(0, grid_depth)
    ax.set_aspect('equal')
    
    # Draw each cell
    for row in range(grid_depth):
        for col in range(grid_width):
            color = grid_colors[row, col]
            rect = patches.Rectangle(
                (col, grid_depth - row - 1),  # Flip vertically
                1, 1,
                facecolor=color,
                edgecolor='white',
                linewidth=0.5
            )
            ax.add_patch(rect)
    
    # Remove axes
    ax.set_xticks([])
    ax.set_yticks([])
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    
    plt.tight_layout()
    return fig

def rgb_to_hex(rgb):
    """Convert RGB values (0-1 range) to hex color code"""
    return '#{:02x}{:02x}{:02x}'.format(
        int(rgb[0] * 255),
        int(rgb[1] * 255),
        int(rgb[2] * 255)
    )

def analyze_carpet_with_claude(image, color_palette):
    """
    Use Claude API to analyze the carpet pattern
    """
    try:
        # Get API key from secrets
        api_key = st.secrets["ANTHROPIC_API_KEY"]
        client = anthropic.Anthropic(api_key=api_key)
        
        # Convert image to base64
        buffered = BytesIO()
        image.save(buffered, format="PNG")
        img_base64 = base64.b64encode(buffered.getvalue()).decode()
        
        # Create color palette text
        palette_text = "\n".join([
            f"Color {i+1}: {rgb_to_hex(color)} (RGB: {int(color[0]*255)}, {int(color[1]*255)}, {int(color[2]*255)})"
            for i, color in enumerate(color_palette)
        ])
        
        # Call Claude API
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1500,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": "image/png",
                                "data": img_base64,
                            },
                        },
                        {
                            "type": "text",
                            "text": f"""Analyze this carpet/rug pattern. The extracted color palette is:

{palette_text}

Please provide:
1. **Pattern Style**: Identify the design style (e.g., Persian, Turkish, Modern, Geometric, Floral, Tribal)
2. **Design Elements**: Describe the key visual elements and motifs
3. **Color Harmony**: Analyze the color scheme and how colors work together
4. **Overall Impression**: Your professional assessment of the design

Keep your response clear, concise, and organized."""
                        }
                    ]
                }
            ]
        )
        
        return message.content[0].text
        
    except Exception as e:
        return f"Error analyzing image: {str(e)}"

def export_grid_to_png(fig):
    """
    Export the grid figure to PNG bytes
    """
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=300, bbox_inches='tight')
    buf.seek(0)
    return buf.getvalue()

def export_grid_to_excel(grid_colors):
    """
    Export the grid to Excel with colors embedded in cells
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Carpet Pattern Grid"
    
    grid_depth, grid_width = grid_colors.shape[:2]
    
    # Set column widths and row heights
    for col in range(1, grid_width + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    for row in range(1, grid_depth + 1):
        ws.row_dimensions[row].height = 60
    
    # Fill cells with colors
    for row in range(grid_depth):
        for col in range(grid_width):
            cell = ws.cell(row=row+1, column=col+1)
            
            # Get color
            color_rgb = grid_colors[row, col]
            hex_color = rgb_to_hex(color_rgb)[1:]  # Remove # symbol
            
            # Set cell background color
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            
            # Add hex code as text
            cell.value = f"#{hex_color}"
            
            # Set text color (white or black based on brightness)
            brightness = (color_rgb[0] * 299 + color_rgb[1] * 587 + color_rgb[2] * 114) / 1000
            text_color = "000000" if brightness > 0.5 else "FFFFFF"
            cell.font = Font(color=text_color, bold=True, size=9)
            
            # Center align
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

# Title
st.title("🎨 Carpet Pattern Analyzer")
st.markdown("Upload a carpet image to extract its pattern grid and color palette")

# Sidebar for controls
with st.sidebar:
    st.header("Controls")
    
    # Grid dimensions
    st.subheader("Grid Dimensions")
    grid_width = st.number_input("Width (cells)", min_value=2, max_value=100, value=10)
    grid_depth = st.number_input("Depth (cells)", min_value=2, max_value=100, value=10)
    
    # Color variation slider (we'll use this in the next step)
    st.subheader("Color Settings")
    color_variation = st.slider(
        "Color Variation",
        min_value=3,
        max_value=20,
        value=10,
        help="Lower = fewer distinct colors, Higher = more color detail"
    )
    
    st.markdown("---")
    st.caption("Step 5: Export features active ✅")

# Main content area
col1, col2 = st.columns(2)

with col1:
    st.subheader("Upload Carpet Image")
    uploaded_file = st.file_uploader(
        "Choose an image...",
        type=['png', 'jpg', 'jpeg', 'webp']
    )
    
    if uploaded_file is not None:
        # Display uploaded image
        image = Image.open(uploaded_file)
        st.image(image, caption="Original Carpet Image", use_container_width=True)
        
        # Show image info
        st.info(f"Image size: {image.size[0]} × {image.size[1]} pixels")

with col2:
    st.subheader("Pattern Grid")
    if uploaded_file is not None:
        with st.spinner("Generating pattern grid..."):
            # Extract colors from grid
            raw_grid_colors = extract_grid_colors(image, grid_width, grid_depth)
            
            # Quantize colors based on slider
            grid_colors, color_palette = quantize_colors(raw_grid_colors, color_variation)
            
            # Plot the grid
            fig = plot_color_grid(grid_colors)
            st.pyplot(fig)
            
            st.success(f"✅ Generated {grid_width} × {grid_depth} grid with {color_variation} distinct colors")
            
            # Export buttons
            st.markdown("---")
            st.subheader("📥 Export Options")
            
            col_export1, col_export2 = st.columns(2)
            
            with col_export1:
                # PNG export
                png_bytes = export_grid_to_png(fig)
                st.download_button(
                    label="⬇️ Download Grid (PNG)",
                    data=png_bytes,
                    file_name="carpet_pattern_grid.png",
                    mime="image/png",
                    use_container_width=True
                )
            
            with col_export2:
                # Excel export
                excel_bytes = export_grid_to_excel(grid_colors)
                st.download_button(
                    label="📊 Download Grid (Excel)",
                    data=excel_bytes,
                    file_name="carpet_pattern_grid.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            st.caption("PNG: High-resolution image | Excel: Editable grid with color codes")
    else:
        st.warning("Upload an image to see the pattern grid")

# AI Analysis section (full width below the two columns)
if uploaded_file is not None:
    st.markdown("---")
    st.subheader("🤖 AI Pattern Analysis")
    
    # Add analyze button
    if st.button("🔍 Analyze Pattern with Claude AI", type="primary", use_container_width=True):
        with st.spinner("Claude is analyzing your carpet pattern..."):
            analysis = analyze_carpet_with_claude(image, color_palette)
            st.session_state['analysis'] = analysis
    
    # Display analysis if it exists
    if 'analysis' in st.session_state:
        st.markdown(st.session_state['analysis'])

# Color palette section (below both columns)
if uploaded_file is not None:
    st.markdown("---")
    st.subheader("🎨 Color Palette")
    
    st.write(f"**{len(color_palette)} distinct colors** in this pattern:")
    
    # Display color palette
    cols_per_row = 10
    num_rows = (len(color_palette) + cols_per_row - 1) // cols_per_row
    
    for row_idx in range(num_rows):
        color_cols = st.columns(cols_per_row)
        start_idx = row_idx * cols_per_row
        end_idx = min(start_idx + cols_per_row, len(color_palette))
        
        for col_idx, color_idx in enumerate(range(start_idx, end_idx)):
            with color_cols[col_idx]:
                color = color_palette[color_idx]
                hex_color = rgb_to_hex(color)
                
                # Color swatch
                st.markdown(
                    f'<div style="background-color: {hex_color}; width: 100%; height: 60px; border: 2px solid #ddd; border-radius: 8px; margin-bottom: 5px;"></div>',
                    unsafe_allow_html=True
                )
                
                # Hex code
                st.markdown(f"**{hex_color}**")
                
                # RGB values
                st.caption(f"RGB: {int(color[0]*255)}, {int(color[1]*255)}, {int(color[2]*255)}")
    
    # Download color list
    st.markdown("---")
    
    # Create downloadable color list
    color_list_text = "CARPET PATTERN COLOR PALETTE\n"
    color_list_text += "=" * 40 + "\n\n"
    
    for idx, color in enumerate(color_palette):
        hex_color = rgb_to_hex(color)
        rgb_values = f"RGB({int(color[0]*255)}, {int(color[1]*255)}, {int(color[2]*255)})"
        color_list_text += f"Color {idx+1}: {hex_color} | {rgb_values}\n"
    
    st.download_button(
        label="📥 Download Color List",
        data=color_list_text,
        file_name="carpet_colors.txt",
        mime="text/plain"
    )